import heapq

import openpyxl

import random
import timeit

import networkx as nx


import numpy as np
from typing import List, Tuple
from heapq import heappop, heappush


def random_connected_graph(n, x, w):

    # Generate a random positive connected graph
    G = nx.fast_gnp_random_graph(n, x)
    while not nx.is_connected(G):
        G = nx.fast_gnp_random_graph(n, x)
    # Assign random weights to the edges
    for (u, v) in G.edges():
        G[u][v]['weight'] = random.randint(w[0], w[1])
    # Create an adjacency matrix with the weights
    adj_matrix = np.zeros((n, n))
    for (u, v, data) in G.edges(data=True):
        adj_matrix[u][v] = data['weight']
        adj_matrix[v][u] = data['weight']
    return adj_matrix


def a_star_search(adj_matrix: np.ndarray, start_node: int, end_node: int) -> Tuple[List[int], int]:
    num_nodes = len(adj_matrix)

    # Initialize the start and goal nodes
    start_node_index = start_node - 1  # adjust to 0-based indexing
    end_node_index = end_node - 1  # adjust to 0-based indexing
    g_score = np.full(num_nodes, np.inf)
    g_score[start_node_index] = 0
    f_score = np.full(num_nodes, np.inf)
    f_score[start_node_index] = euclidean_distance(adj_matrix[start_node_index], adj_matrix[end_node_index])

    # Initialize the set of explored nodes and the priority queue
    explored_set = set()
    priority_queue = [(f_score[start_node_index], start_node_index)]

    # Initialize the came_from dictionary with the start node
    came_from = {start_node_index: -1}

    # Run the A* algorithm
    while priority_queue:
        # Get the node with the lowest f-score from the priority queue
        _, curr_node_index = heappop(priority_queue)

        # If the current node is the goal node, return the path
        if curr_node_index == end_node_index:
            path = reconstruct_path(came_from, end_node_index)
            return path, g_score[end_node_index]

        # Add the current node to the explored set
        explored_set.add(curr_node_index)

        # Explore the neighbors of the current node
        for neighbor_index, weight in enumerate(adj_matrix[curr_node_index]):
            if weight > 0:
                if neighbor_index not in explored_set:
                    tentative_g_score = g_score[curr_node_index] + weight
                    if tentative_g_score < g_score[neighbor_index]:
                        # This path to neighbor is better than any previous one. Record it!
                        came_from[neighbor_index] = curr_node_index
                        g_score[neighbor_index] = tentative_g_score
                        f_score[neighbor_index] = tentative_g_score + euclidean_distance(
                            adj_matrix[neighbor_index], adj_matrix[end_node_index])
                        if neighbor_index not in priority_queue:
                            heappush(priority_queue, (f_score[neighbor_index], neighbor_index))

    # If the goal node cannot be reached, return an empty path and infinity distance
    return [], np.inf


def euclidean_distance(a: np.ndarray, b: np.ndarray) -> float:
    return np.sqrt(np.sum((a - b) ** 2))


def reconstruct_path(came_from: np.ndarray, end_node_index: int) -> List[int]:
    path = [end_node_index]
    while came_from[path[0]] >= 0:
        path.insert(0, came_from[path[0]])
    return path



def dijkstra(dist_matrix, start_node, end_node):
    n = len(dist_matrix)  # number of nodes
    pq = [(0, start_node)]  # priority queue with start node
    visited = {}  # visited nodes and their distances
    parent = {}  # parent node of each visited node

    # initialize distances to infinity except for start node
    for i in range(n):
        visited[i] = float('inf')
    visited[start_node] = 0

    while pq:
        # get node with smallest distance from start node
        (dist, curr_node) = heapq.heappop(pq)

        # check if we've reached the end node
        if curr_node == end_node:
            path = []
            node = curr_node
            while node in parent:
                path.append(node)
                node = parent[node]
            path.append(start_node)
            path.reverse()
            return (path, dist)

        # update distances of neighbors
        for neighbor, weight in enumerate(dist_matrix[curr_node]):
            if weight > 0:  # if edge exists
                new_dist = dist + weight
                if new_dist < visited[neighbor]:
                    visited[neighbor] = new_dist
                    parent[neighbor] = curr_node
                    heapq.heappush(pq, (new_dist, neighbor))

    # end node not reachable from start node
    return (None, None)

# Define the range of graph sizes
graph_sizes = range(100, 6000, 100)

# Initialize empty lists to store the average time it takes to execute each algorithm
avg_astar_times = []
avg_dijkstra_times = []



    # Iterate over the graph sizes and calculate the time it takes to execute each algorithm

t = 0
for n in graph_sizes:

    astar_times = []
    dijkstra_times = []
    print("graph size:", n)
    ll = 0
    for run in range(50):
        print(run)
        rn= random.randint(0, n-1)  # generate a random integer between 1 and n
        x = 0.04  # Specify the probability of two nodes being connected
        w = (1, 10)
        adj_matrix = random_connected_graph(n, x, w)

        start_time1 = timeit.default_timer()
        path, weight = a_star_search(adj_matrix, 0, rn)
        astar_time = timeit.default_timer() - start_time1
        astar_times.insert(ll, astar_time)

        start_time2 = timeit.default_timer()
        z = dijkstra(adj_matrix,0, rn)
        dijkstra_time = timeit.default_timer() - start_time2
        dijkstra_times.insert(ll, dijkstra_time)
        ll = ll + 1
    d = np.mean(dijkstra_times)
    a = np.mean(astar_times)
    avg_dijkstra_times.insert(t, d)
    avg_astar_times.insert(t, a)
    t = t + 1

# Print the final results
print("Final results:")
print(f"Average A* times: {avg_astar_times}")
print(f"Average Dijkstra's times: {avg_dijkstra_times}")
i = 0
workbook = openpyxl.Workbook()
worksheet = workbook.active
for i in range(len(avg_astar_times)):
    worksheet.cell(row=i+1, column=1, value=avg_astar_times[i])
for i in range(len(avg_dijkstra_times)):
    worksheet.cell(row=i+1, column=2, value=avg_dijkstra_times[i])
workbook.save("Astar VS Dijkstras prob 0.04 TESTTTT.xlsx")
